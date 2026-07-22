from pathlib import Path

import pandas as pd

import engine


def test_money_parses_common_formats():
    assert engine.money("1,234.50") == 1234.5
    assert engine.money("￥ 88.00") == 88.0
    assert engine.money("(20.5)") == -20.5
    assert engine.money(12) == 12.0


def test_read_file_detects_header_after_bank_metadata(tmp_path: Path):
    path = tmp_path / "statement.csv"
    path.write_text("账户名称,测试公司\n查询期间,2026-01\n交易日期,摘要,收入,支出\n2026-01-02,销售回款,1000,\n", encoding="utf-8-sig")
    df = engine.read_file(path)
    assert list(df.columns)[:4] == ["交易日期", "摘要", "收入", "支出"]
    assert len(df) == 1


def test_rule_classification_uses_direction():
    income = engine.rule_classify({"summary": "客户货款回款", "counterparty": "甲客户", "remark": "", "direction": "收入", "signed_amount": 1000})
    expense = engine.rule_classify({"summary": "供应商材料货款", "counterparty": "乙供应商", "remark": "", "direction": "支出", "signed_amount": -500})
    assert income["category"] == "销售商品、提供劳务收到的现金"
    assert expense["category"] == "购买商品、接受劳务支付的现金"


def test_cashflow_statement_totals():
    job = {"results": [
        {"category": "销售商品、提供劳务收到的现金", "ctx": {"signed_amount": 1000}, "review": False},
        {"category": "购买商品、接受劳务支付的现金", "ctx": {"signed_amount": -300}, "review": False},
        {"category": "取得借款收到的现金", "ctx": {"signed_amount": 500}, "review": False},
    ]}
    cf = engine.cashflow_statement(job)
    assert cf["net_increase"] == 1200


def test_excel_export_contains_financial_report_pack(tmp_path: Path):
    from openpyxl import load_workbook
    job = {
        "rows": [
            {"来源文件": "bank-a.csv", "日期": "2026-01-01", "摘要": "销售回款"},
            {"来源文件": "bank-b.csv", "日期": "2026-01-02", "摘要": "房租"},
        ],
        "files": ["bank-a.csv", "bank-b.csv"],
        "results": [
            {"category": "销售商品、提供劳务收到的现金", "reason": "客户回款", "confidence": 0.95,
             "review": False, "source": "规则", "ctx": {"date": "2026-01-01", "summary": "销售回款", "counterparty": "甲客户", "signed_amount": 1000}},
            {"category": "支付其他与经营活动有关的现金", "reason": "办公场地租金", "confidence": 0.88,
             "review": True, "source": "AI", "ctx": {"date": "2026-01-02", "summary": "房租", "counterparty": "乙物业", "signed_amount": -300}},
        ],
    }
    out = engine.export_excel(job, tmp_path / "reports.xlsx")
    sheets = set(load_workbook(out, read_only=True).sheetnames)
    assert {"分类结果", "现金流量表", "资金收支总览", "每日资金收支", "收入分类分析",
            "费用支出分析", "往来单位分析", "待复核流水", "多账户来源汇总"} <= sheets
